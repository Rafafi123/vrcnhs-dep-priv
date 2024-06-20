from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, F, Value as V
from django.db.models.functions import Coalesce
from datetime import datetime
import plotly.graph_objs as go
import plotly.express as px
from django.core.exceptions import PermissionDenied
from django.contrib.auth.views import PasswordChangeView, PasswordChangeDoneView
from dateutil import parser as date_parser
from django.urls import reverse, reverse_lazy
from io import BytesIO
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User, Group
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from openpyxl import load_workbook
from tablib import Dataset
from data.resources import ClassroomResource, StudentResource
from .forms import StudentForm
from .forms import TeacherForm
from .models import Student, Classroom, Teacher, Gradelevel
from django.db.models import Count
from django.views.generic import TemplateView # this is needed for the charts and graphs
from data.models import Student
import plotly.graph_objs as go
import plotly.io as pio
from .forms import StudentSearchForm, StudentForm, TeacherSearchForm, TeacherSignupForm, AdminTeacherStudentForm  # this is for the search function
import pandas as pd # this is for the data analysis
from django.db.models import Count, Avg
from django.http import FileResponse, HttpResponseForbidden, JsonResponse
from django.views.decorators.cache import cache_page #
from chartjs.views.lines import BaseLineChartView 
from django.http import HttpResponse
from django.db.models import Q
from .decorators import unauthenticated_user, allowed_users, admin_only
import random
import plotly.express as px
import datetime
import plotly.graph_objects as go
import logging
from django.forms.models import model_to_dict
from plotly.io import to_html as pio_to_html
# Create your views here.

@unauthenticated_user
def signup(request):
    if request.method == 'POST':
        teacher_form = TeacherSignupForm(request.POST)
        if teacher_form.is_valid():
            user = teacher_form.save()
            messages.success(request, "User registration was successful!")
            return redirect('login')
    else:
        teacher_form = TeacherSignupForm()
    return render(request, 'registration/signup.html', {
        'teacher_form': teacher_form
    })

#OG Login, signup pages
def create_pie_chart(labels, sizes, title, chart_width=None, chart_height=None):
    fig = go.Figure(data=[go.Pie(labels=labels, values=sizes)])
    fig.update_layout(title=title, autosize=True, width=chart_width, height=chart_height)
    return fig

def create_bar_chart(labels, sizes, title, chart_width=None, chart_height=None, colorscale='bright'):
    colors = px.colors.qualitative.Plotly
    fig = go.Figure(data=[go.Bar(x=labels, y=sizes, marker_color=colors)])
    
    fig.update_xaxes(fixedrange=True)
    fig.update_yaxes(fixedrange=True)
    fig.update_layout(title=title, autosize=True, width=chart_width, height=chart_height)

    fig.update_layout(
        autosize=True,
        margin=dict(l=0, r=0, b=0, t=30),
        template="plotly",
    )

    return fig

@login_required(login_url='login')
def home(request):
    if request.user.groups.filter(name='ADMIN').exists():
        # Logic for ADMIN users
        count = User.objects.count()
        
        # Calculate male and female student counts
        male_count = Student.objects.filter(Q(sex='M') | Q(sex='Male')).count()
        female_count = Student.objects.filter(Q(sex='F') | Q(sex='Female')).count()

        # Calculate total student count
        total_students = male_count + female_count

        # Create pie chart for gender distribution
        gender_fig = create_pie_chart(
            labels=['Male', 'Female'],
            sizes=[male_count, female_count],
            title='Student Gender Distribution'
        )
        gender_chart_div = pio_to_html(gender_fig, full_html=False, include_plotlyjs='cdn')

        students = Student.objects.all()

        # Calculate religion distribution
        religion_counts = students.values(religion_value=Coalesce('religion', V('None'))).annotate(count=Count('religion_value'))

        religion_fig = create_bar_chart(
            [item['religion_value'] for item in religion_counts],
            [item['count'] for item in religion_counts],
            'Distribution of Religions'
        )

        # Calculate scholarship distribution
        scholarship_counts = students.values(scholarship_value=Coalesce('is_a_four_ps_scholar', V('None'))).annotate(count=Count('scholarship_value'))

        scholarship_fig = create_bar_chart(
            [item['scholarship_value'] for item in scholarship_counts],
            [item['count'] for item in scholarship_counts],
            'Distribution of Scholars'
        )

        # Get the current date and time
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Retrieve total teachers and classrooms
        total_teachers = Teacher.objects.count()
        total_classrooms = Classroom.objects.count()

        # Add variables to context dictionary
        context = {
            'total_students': total_students,
            'gender_chart_div': gender_chart_div,
            'religion_chart': religion_fig.to_html(full_html=False, include_plotlyjs='cdn'),
            'current_datetime': current_datetime,
            'count': count,
            'total_teachers': total_teachers,
            'total_classrooms': total_classrooms,
            'scholarship_chart': scholarship_fig.to_html(full_html=False, include_plotlyjs='cdn'),
        }

        # Render the home.html template with the context data
        return render(request, 'home.html', context)
    
    elif request.user.groups.filter(name='TEACHER').exists():
        if request.user.groups.filter(name='ADMIN').exists():
            # Logic for users in both 'TEACHER' and 'ADMIN' groups
            # Redirect to the home page
            return render(request, 'home.html', context)
        else:
            # Logic for users only in 'TEACHER' group
            # Redirect to the user_page for TEACHER users
            return redirect('user_page')
    

################################################## for class organization

@login_required
def grade_sections(request):
    classrooms_grade_7= Classroom.objects.filter(gradelevel__grade='Grade 7')
    classrooms_grade_8= Classroom.objects.filter(gradelevel__grade='Grade 8')
    classrooms_grade_9= Classroom.objects.filter(gradelevel__grade='Grade 9')
    classrooms_grade_10= Classroom.objects.filter(gradelevel__grade='Grade 10')
    classrooms_grade_11= Classroom.objects.filter(gradelevel__grade='Grade 11')
    classrooms_grade_12= Classroom.objects.filter(gradelevel__grade='Grade 12')

    context = {
        'classrooms_grade_7': classrooms_grade_7,
        'classrooms_grade_8': classrooms_grade_8,
        'classrooms_grade_9': classrooms_grade_9,
        'classrooms_grade_10': classrooms_grade_10,	
        'classrooms_grade_11': classrooms_grade_11,
        'classrooms_grade_12': classrooms_grade_12,
        }

    return render(request, 'grade_sections.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['TEACHER'])
def user_page(request):
    # Retrieve all classrooms associated with the teacher
    classrooms = request.user.teacher.classroom_set.all()
    
    if classrooms.exists():
        # If at least one classroom exists
        classroom = classrooms[0]  # Get the first classroom
        students = Student.objects.filter(classroom=classroom)  # Filter students based on the classroom
        teacher_name = f"{request.user.teacher.first_name} {request.user.teacher.last_name}"
    else:
        # If no classrooms exist
        classroom = None  # Set classroom to None
        students = []  # Set students as an empty list
        teacher_name = None  # Set teacher_name to None

    # Check if the teacher is not assigned a classroom
    if request.user.teacher.classroom_set.count() == 0:
        classroom_name = None
    else:
        classroom_name = classroom.gradelevel.grade + ' "' + classroom.classroom + '"'

    context = {'classroom_name': classroom_name, 'teacher_name': teacher_name, 'students': students}
    return render(request, 'user_page.html', context)

@login_required
def classroom_detail(request, classroom_id): # this is for the individual classrooms selected which will bring the user to the class list
    classroom = get_object_or_404(Classroom, pk=classroom_id)
    students = Student.objects.filter(classroom=classroom).order_by('last_name', 'first_name')
    context = {'classroom': classroom, 'students': students}
    
    return render(request, 'classroom_detail.html', context)
    

#####################################################################
@login_required
@allowed_users(allowed_roles=['ADMIN', 'TEACHER'])
def edit_teacher(request, teacher_id):
    teacher = Teacher.objects.get(id=teacher_id)
    user = teacher.user

    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=teacher)
        username = request.POST.get('username', '')

        if form.is_valid() and not User.objects.filter(username=username).exclude(pk=user.pk).exists():
            user.username = username
            user.save()
            form.save()
            return redirect('teachers')
        else:
            messages.error(request, "Username already exists or form is invalid.")
    else:
        form = TeacherForm(instance=teacher)

    return render(request, 'edit_teacher.html', {'form': form, 'teacher_id': teacher_id, 'username': user.username})

@login_required
def students_page(request):
        student_list = Student.objects.all()
        return render(request,"view_students.html",{"students":student_list}) # this is to show the objects in the database


################################################################################ The CRUD Functions 
@login_required
def add_student(request):
    user = request.user

    if user.groups.filter(name='TEACHER').exists():
        teacher = user.teacher
        is_admin = user.groups.filter(name='ADMIN').exists()

        if request.method == 'POST':
            form = StudentForm(request.POST, teacher=teacher, is_admin=is_admin)
            if form.is_valid():
                form.save()
                print(f"Student added: {form.instance}")  # Print the saved student instance
                messages.success(request, "Student Added Successfully") # Notify user of successful Creation of student
                return redirect("user_page")
        else:
            form = StudentForm(teacher=teacher, is_admin=is_admin)
            form.fields['classroom'].queryset = Classroom.objects.filter(teacher=teacher)

    elif user.groups.filter(name='ADMIN').exists():
        if request.method == 'POST':
            form = AdminTeacherStudentForm(request.POST)
            if form.is_valid():
                form.save()
                print(f"Student added: {form.instance}")  # Print the saved student instance
                messages.success(request, "Student Added Successfully") # Notify user of successful Creation of student
                return redirect("students")
        else:
            form = StudentForm()
    else:
        # Handle other user groups or unauthorized access as desired
        return HttpResponse("Unauthorized access")

    context = {'form': form}
    return render(request, 'add_student.html', context)

#DELETE TEACHER/USER
@login_required
@allowed_users(allowed_roles=['ADMIN', 'TEACHER'])
def destroy_teacher(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)

    # Check if the user has permission to delete the teacher
    if request.user.has_perm('delete_teacher', teacher):
        teacher.delete()
        messages.error(request, "Teacher Deleted", extra_tags='danger')
        return redirect("teachers")
    else:
        # Handle unauthorized access (optional)
        return HttpResponse("Unauthorized access")  # Create an unauthorized_access.html template
    

#DELETE STUDENT
@login_required
def destroy(request, lrn):
    student = Student.objects.get(LRN=lrn)
    student.delete()
    messages.error(request, "Student Deleted", extra_tags='danger')  # Set extra_tags to 'danger' for red color
    
    # Check if the user belongs to the "TEACHER" group
    if request.user.groups.filter(name='TEACHER').exists():
        # Redirect to "user_page" if the user is a teacher
        return redirect("user_page")
    else:
        # Redirect to "students" if the user is not a teacher
        return redirect("students")

@login_required
def edit(request, lrn):
    student = get_object_or_404(Student, LRN=lrn)
    initial_data = model_to_dict(student)  # Store initial data for comparison
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            edited_fields = []  # Create a list to store edited fields and their previous values

            # Compare the new data with the initial data to detect edits
            for field_name, new_value in form.cleaned_data.items():
                if initial_data[field_name] != new_value:
                    previous_value = initial_data[field_name]
                    field_verbose_name = Student._meta.get_field(field_name).verbose_name
                    edited_fields.append(f"{field_verbose_name} (before: {previous_value})")

            # Store edited fields in the student object
            student.edited_fields = ', '.join(edited_fields)
            student.save()

            form.save()
            messages.success(request, "Student Updated Successfully")

            if request.user.groups.filter(name='TEACHER').exists():
                return redirect("view_student_detail", lrn=lrn)
            else:
                return redirect("view_student_detail", lrn=lrn)
    else:
        # Create the form instance without providing initial values
        form = StudentForm(instance=student)

        # Manually set the initial value for the 'sex' field
        form.fields['sex'].initial = student.sex
    context = {'form': form, 'student': student}
    return render(request, 'edit.html', context)

#VIEW STUDENTS DETAIL

@login_required
def view_student_detail(request, lrn):
    user = request.user
    has_authorization = False

    try:
        # Check if the user is in the ADMIN group
        if Group.objects.get(name='ADMIN') in user.groups.all():
            # If the user is in the ADMIN group, allow access to all student details
            student = Student.objects.get(LRN=lrn)
            has_authorization = True
        else:
            # If the user is not in the ADMIN group, assume they are a teacher
            teacher = Teacher.objects.get(user=user)
            classroom = Classroom.objects.get(teacher=teacher)

            # Try to get the student only if they belong to the same classroom as the teacher
            student = Student.objects.get(LRN=lrn, classroom=classroom)
            has_authorization = True

    except (Group.DoesNotExist, Teacher.DoesNotExist, Classroom.DoesNotExist, Student.DoesNotExist):
        # Handle the case where the user is not a teacher, is not associated with a classroom,
        # or the student is not found. Redirect to the "students" page with an error message.
        messages.error(request, "You are not authorized to view this student's profile.")
        return redirect("students")

    context = {'student': student, 'has_authorization': has_authorization}
    return render(request, 'view_student_detail.html', context)

def has_authorization(user, lrn):
    # Check if the user is in the ADMIN group
    if Group.objects.get(name='ADMIN') in user.groups.all():
        return True
    else:
        # If the user is not in the ADMIN group, assume they are a teacher
        try:
            teacher = Teacher.objects.get(user=user)
            classroom = Classroom.objects.get(teacher=teacher)

            # Try to get the student only if they belong to the same classroom as the teacher
            student = get_object_or_404(Student, LRN=lrn, classroom=classroom)
            return True
        except (Teacher.DoesNotExist, Classroom.DoesNotExist, Student.DoesNotExist):
            # Handle the case where the user is not a teacher, is not associated with a classroom,
            # or the student is not found
            return False

@login_required
def students_page(request):
    student_list = Student.objects.all()
    
    # Check for authorization logic
    user = request.user
    has_authorization_status = has_authorization(user)

    context = {
        'students': student_list,
        'has_authorization': has_authorization_status,
    }

    return render(request, 'view_students.html', context)

def back_student_detail(request):
    user = request.user
    if Group.objects.get(name='TEACHER') in user.groups.all():
        # Redirect to teacher's user_page.html
        return redirect('user_page')  
    else:
        return redirect('students')

@login_required(login_url='login')
def update(request, id):
    student = Student.objects.get(id=id)
    form = StudentForm(instance=student)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            print("Debug Statement: Student Updated -", student.name)  # Debug statement
            messages.success(request, "Student Updated Successfully") # Notify user of successful UPDATES of student
            return redirect("students")
    context = {'form': form, 'student': student}
    return render(request, 'update.html', context)



#######################################################################################

# this is the search function
@login_required(login_url='login')
def students_page(request):
    form = StudentSearchForm(request.GET or None)
    students = Student.objects.all().order_by('last_name', 'first_name')

    if form.is_valid():
        query = form.cleaned_data['query']
        students = students.filter(
            Q(LRN__icontains=query) |
            Q(last_name__icontains=query) |
            Q(first_name__icontains=query) |
            Q(middle_name__icontains=query) |
            Q(suffix_name__icontains=query) |
            Q(classroom__classroom__icontains=query)   # Filter based on the 'classroom' field in the Classroom model
        ).order_by('last_name', 'first_name')
        print("Debug Statement: Searched Student -", query)  # Debug statement
        
    return render(request, 'view_students.html', {'form': form, 'students': students})

@login_required(login_url='login')
@allowed_users(allowed_roles=['ADMIN','TEACHER'])
def teachers_page(request):
    form = TeacherSearchForm(request.GET or None)
    teachers = Teacher.objects.all()

    if form.is_valid():
        query = form.cleaned_data['query']
        teachers = teachers.filter(
            Q(last_name__icontains=query) |
            Q(first_name__icontains=query)
        )
        print("Debug Statement: Searched Teacher -", query)  # Debug statement
        
    context = {'form': form, 'teachers': teachers}
    return render(request, 'view_teachers.html', context)




#################################### this to show the teachers data on the teachers page
#@login_required
#def teachers_page(request):
 #   teachers = Teacher.objects.all()
  #  context = {'teachers': teachers}
   # return render(request, 'view_teachers.html', context)

@login_required
def add_teacher(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('teachers_page')
    else:
        form = TeacherForm()
    context = {'form': form}
    return render(request, 'add_teacher.html', context)

############################## This is for the report page ###########################################

def create_pie_chart(labels, sizes, title, chart_width=None, chart_height=None):
    fig = go.Figure(data=[go.Pie(labels=labels, values=sizes)])
    fig.update_layout(title=title, autosize=True, width=chart_width, height=chart_height)
    return fig

def create_bar_chart(labels, sizes, title, chart_width=None, chart_height=None, colorscale='bright'):
    colors = px.colors.qualitative.Plotly
    fig = go.Figure(data=[go.Bar(x=labels, y=sizes, marker_color=colors)])
    
    fig.update_xaxes(fixedrange=True)
    fig.update_yaxes(fixedrange=True)
    fig.update_layout(title=title, autosize=True, width=chart_width, height=chart_height)

    fig.update_layout(
        autosize=True,
        margin=dict(l=0, r=0, b=0, t=30),
        template="plotly",
    )

    return fig

@login_required
def report_page(request):
    selected_gradelevel = request.GET.get('gradelevel')
    gradelevels = Gradelevel.objects.all()
    user_is_teacher = request.user.groups.filter(name='TEACHER').exists()
    user_is_admin = request.user.groups.filter(name='ADMIN').exists()

    # Default to 'all' for admin users if no grade level is selected
    if user_is_admin and not selected_gradelevel:
        selected_gradelevel = 'all'
    elif user_is_teacher and not selected_gradelevel:
        selected_gradelevel = 'myclassroom'

    if user_is_teacher:
        teacher_classrooms = request.user.teacher.classroom_set.all()
        if selected_gradelevel == 'myclassroom':
            students = Student.objects.filter(classroom__in=teacher_classrooms).select_related('gradelevel')
        elif selected_gradelevel == 'all':
            students = Student.objects.all().select_related('gradelevel')
        else:
            try:
                selected_gradelevel = int(selected_gradelevel)
                students = Student.objects.filter(gradelevel_id=selected_gradelevel).select_related('gradelevel')
            except ValueError:
                students = Student.objects.none()
    elif user_is_admin:
        if selected_gradelevel == 'all':
            students = Student.objects.all().select_related('gradelevel')
        else:
            try:
                selected_gradelevel = int(selected_gradelevel)
                students = Student.objects.filter(gradelevel_id=selected_gradelevel).select_related('gradelevel')
            except ValueError:
                students = Student.objects.none()
    else:
        students = Student.objects.all().select_related('gradelevel')

    # Aggregate counts using Django ORM, ensuring to give unique names to the annotations
    strand_counts = students.values(strand_value=Coalesce('strand', V('None'))).annotate(count=Count('strand_value'))
    economic_counts = students.values(household_income_value=Coalesce('household_income', V('None'))).annotate(count=Count('household_income_value'))
    religion_counts = students.values(religion_value=Coalesce('religion', V('None'))).annotate(count=Count('religion_value'))
    dropout_counts = students.values(is_a_dropout_value=Coalesce('is_a_dropout', V('None'))).annotate(count=Count('is_a_dropout_value'))
    working_student_counts = students.values(is_a_working_student_value=Coalesce('is_a_working_student', V('None'))).annotate(count=Count('is_a_working_student_value'))
    scholarship_counts = students.values(is_a_four_ps_scholar_value=Coalesce('is_a_four_ps_scholar', V('None'))).annotate(count=Count('is_a_four_ps_scholar_value'))
    sex_counts = students.values(sex_value=Coalesce('sex', V('None'))).annotate(count=Count('sex_value'))
    returnee_counts = students.values(is_returnee_value=Coalesce('is_returnee', V('None'))).annotate(count=Count('is_returnee_value'))

    # Create charts
    strand_fig = create_bar_chart(
        [item['strand_value'] for item in strand_counts],
        [item['count'] for item in strand_counts],
        'Distribution of Academic Strands'
    )
    
    economic_fig = create_bar_chart(
        [item['household_income_value'] for item in economic_counts],
        [item['count'] for item in economic_counts],
        'Distribution of Household Income'
    )

    religion_fig = create_bar_chart(
        [item['religion_value'] for item in religion_counts],
        [item['count'] for item in religion_counts],
        'Distribution of Religions'
    )

    dropout_fig = create_pie_chart(
        [item['is_a_dropout_value'] for item in dropout_counts],
        [item['count'] for item in dropout_counts],
        'Distribution of Dropout Status'
    )

    working_student_fig = create_pie_chart(
        [item['is_a_working_student_value'] for item in working_student_counts],
        [item['count'] for item in working_student_counts],
        'Distribution of Working Students'
    )

    scholarship_fig = create_bar_chart(
        [item['is_a_four_ps_scholar_value'] for item in scholarship_counts],
        [item['count'] for item in scholarship_counts],
        'Distribution of Scholars'
    )

    sex_fig = create_pie_chart(
        [item['sex_value'] for item in sex_counts],
        [item['count'] for item in sex_counts],
        'Males and Females'
    )

    returnee_fig = create_pie_chart(
        [item['is_returnee_value'] for item in returnee_counts],
        [item['count'] for item in returnee_counts],
        'Distribution of Returnee Students'
    )

    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if selected_gradelevel == 'myclassroom':
        selected_filter_name = "My Classroom"
    elif selected_gradelevel == "all":
        selected_filter_name = "All Grade Levels"
    else:
        try:
            selected_grade_object = Gradelevel.objects.get(id=selected_gradelevel)
            selected_filter_name = f"{selected_grade_object.grade}"
        except Gradelevel.DoesNotExist:
            selected_filter_name = "Unknown Grade Level"

    return render(request, 'report_page.html', {
        'current_datetime': current_datetime,
        'strand_chart': strand_fig.to_html(full_html=False, include_plotlyjs='cdn'),
        'economic_chart': economic_fig.to_html(full_html=False, include_plotlyjs='cdn'),
        'religion_chart': religion_fig.to_html(full_html=False, include_plotlyjs='cdn'),
        'dropout_chart': dropout_fig.to_html(full_html=False, include_plotlyjs='cdn'),
        'working_student_chart': working_student_fig.to_html(full_html=False, include_plotlyjs='cdn'),
        'scholarship_chart': scholarship_fig.to_html(full_html=False, include_plotlyjs='cdn'),
        'sex_chart': sex_fig.to_html(full_html=False, include_plotlyjs='cdn'),
        'returnee_chart': returnee_fig.to_html(full_html=False, include_plotlyjs='cdn'),
        'gradelevels': gradelevels,
        'selected_filter_name': selected_filter_name,
        'user_is_teacher': user_is_teacher,
        'selected_gradelevel': selected_gradelevel,  # Pass the selected grade level to the template
    })

############################### this is for adding classrooms and assigning a teacher to those classrooms
def add_classroom(request):
    if request.method == 'POST':
        # Retrieve the values from the form
        classroom_name = request.POST.get('classroom_name')
        gradelevel_id = request.POST.get('gradelevel_id')
        teacher_id = request.POST.get('teacher_id')

        try:
            # Attempt to retrieve the Gradelevel object with the given ID
            gradelevel = Gradelevel.objects.get(id=gradelevel_id)

            if teacher_id:  # Check if a teacher is selected
                # Retrieve the Teacher object with the given ID
                teacher = Teacher.objects.get(id=teacher_id)
            else:
                teacher = None  # Set teacher to None if no teacher is selected

            # Create a new Classroom object with the provided values
            Classroom.objects.create(classroom=classroom_name, gradelevel=gradelevel, teacher=teacher)
            
            # Debug print statements
            print("Debug Statement: Classroom added")
            print("Classroom name:", classroom_name)
            print("Teacher:", teacher)
            print("Grade level:", gradelevel)
            
            return redirect('grade_sections')
        except Gradelevel.DoesNotExist:
            print("Gradelevel does not exist")  # Debug statement
            # Handle the error or redirect to an appropriate page
        except Teacher.DoesNotExist:
            print("Teacher does not exist")  # Debug statement
            # Handle the error or redirect to an appropriate page

    # Retrieve all the available grade levels and teachers
    gradelevels = Gradelevel.objects.all()
    teachers = Teacher.objects.all()

    
    
    # Render the template with the grade levels and teachers
    return render(request, 'add_classroom.html', {'gradelevels': gradelevels, 'teachers': teachers})



############################### this is for editting the classrooms and their assigned teacher
@login_required(login_url='login')
@allowed_users(allowed_roles=['ADMIN', 'TEACHER'])
def edit_classroom(request, classroom_id):
    try:
        # Retrieve the classroom object based on the provided classroom_id
        classroom = Classroom.objects.get(id=classroom_id)
    except Classroom.DoesNotExist:
        # Handle the case where the classroom doesn't exist
        return HttpResponse("Classroom does not exist.")

    if request.method == 'POST':
        # Retrieve the form data submitted via POST
        classroom_name = request.POST.get('classroom_name')
        gradelevel_id = request.POST.get('gradelevel_id')
        teacher_id = request.POST.get('teacher_id')

        try:
            # Retrieve the gradelevel object based on the selected gradelevel_id
            gradelevel = Gradelevel.objects.get(id=gradelevel_id)
        except Gradelevel.DoesNotExist:
            # Handle the case where the grade level doesn't exist
            return HttpResponse("Grade level does not exist.")

        if teacher_id == "-1":
            # If the selected teacher_id is -1 (None), set the teacher to None
            teacher = None
        else:
            try:
                # Retrieve the teacher object based on the selected teacher_id
                teacher = Teacher.objects.get(id=teacher_id)
            except Teacher.DoesNotExist:
                # Handle the case where the teacher doesn't exist
                return HttpResponse("Teacher does not exist.")

        # Update the classroom object with the new data
        classroom.classroom = classroom_name
        classroom.gradelevel = gradelevel
        classroom.teacher = teacher
        classroom.save()
        
        # Build the URL for the current view (edit_classroom) without the success message
        redirect_url = reverse(grade_sections)
        
        # Add the success message to the messages framework
        success_message = f'Classroom was successfully edited! You may go back and press refresh to see the changes.'
        messages.success(request, success_message)

        # Redirect to the appropriate page after saving
        return redirect('edit_classroom', classroom_id=classroom_id)

    # Retrieve all gradelevels and teachers for rendering the form
    gradelevels = Gradelevel.objects.all()
    teachers = Teacher.objects.all()

    # Prepare the context to pass to the template
    context = {
        'classroom': classroom,
        'gradelevels': gradelevels,
        'teachers': teachers,
        'current_gradelevel': classroom.gradelevel.id
    }

    print("Debug Statement: Edited Classroom")
    print("Classroom name:", classroom.classroom)
    print("Teacher:", classroom.teacher)
    print("Grade level:", classroom.gradelevel)

    # Render the edit_classroom.html template with the provided context
    return render(request, 'edit_classroom.html', context)




############################### this is for deleting classrooms
@login_required(login_url='login')
@allowed_users(allowed_roles=['ADMIN'])
def delete_classroom(request, classroom_id):
    classroom = get_object_or_404(Classroom, id=classroom_id)

    if request.method == 'POST':
        print("Debug Statement: Deleting Classroom -", classroom.classroom)  # Debug statement
        classroom.delete()
        
        teachers = Teacher.objects.all()  # Add this line to retrieve all teachers
    return redirect('grade_sections')

########################### STUDENT RECORD
@login_required(login_url='login')
def student_record(request):
    students = Student.history.all()  # Retrieve all historical records
    history_data = []

    for record in students:
        changes = record.instance.get_changes()  # Use instance to call get_changes on the actual student instance
        if changes:
            history_data.append({
                'student': record.instance,
                'changes': changes,
                'history_date': record.history_date,
                'history_user': record.history_user,
                'history_type': record.history_type,
            })

    context = {'history_data': history_data}
    return render(request, 'student_record.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['ADMIN'])
def delete_all_student_history(request):
    if request.method == 'POST':
        Student.history.all().delete()
        messages.success(request, "All historical records have been deleted.")
        return redirect('student_record')

    return render(request, 'confirm_delete.html')

#TEMPLATE
def download_template(request):
    # Logic to serve the download link for the template file
    existing_wb = load_workbook('data/static/media/VRCNHS_STUDENT_TEMPLATE(CLEAR).xlsx')

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    existing_wb.save(buffer)
    buffer.seek(0)

    response = FileResponse(buffer, as_attachment=True, filename='Students_Template.xlsx')
    
    return response


#EXPORT student

@allowed_users(allowed_roles=['ADMIN', 'TEACHER'])
def export_students_to_excel(request):
    try:
        students = Student.objects.all()

        # Load the template workbook
        template_path = 'data/static/media/VRCNHS_STUDENT_TEMPLATE.xlsx'
        existing_wb = load_workbook(template_path)
        sheet = existing_wb.active

        # Clear existing data
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

        start_row = 2
        start_column = 2

        date_style = NamedStyle(name='date_style', number_format='MM-DD-YYYY')

        # Define a list representing the order of columns in the Excel file
        excel_columns_order = [
            'LRN', 'last_name', 'first_name', 'middle_name', 'suffix_name', 'status', 'birthday',
            'religion', 'other_religion', 'strand', 'age', 'sem', 'classroom', 'gradelevel', 'sex',
            'birth_place', 'mother_tongue', 'address', 'father_name', 'father_contact', 'mother_name',
            'mother_contact', 'guardian_name', 'guardian_contact', 'transfer_status', 'household_income',
            'is_returnee', 'is_a_dropout', 'is_a_working_student', 'health_bmi', 'general_average',
            'is_a_four_ps_scholar', 'notes',
            '',  # Blank column
            '',
            # Grade 7
            'g7_school', 'g7_schoolYear', 'g7_section', 'g7_general_average', 'g7_adviser', 'g7_adviserContact',
            '',  # Blank column
            # Grade 8
            'g8_school', 'g8_schoolYear', 'g8_section', 'g8_general_average', 'g8_adviser', 'g8_adviserContact',
            '',  # Blank column
            # Grade 9
            'g9_school', 'g9_schoolYear', 'g9_section', 'g9_general_average', 'g9_adviser', 'g9_adviserContact',
            '',  # Blank column
            # Grade 10
            'g10_school', 'g10_schoolYear', 'g10_section', 'g10_general_average', 'g10_adviser', 'g10_adviserContact',
            '',  # Blank column
            # Grade 11
            'g11_school', 'g11_schoolYear', 'g11_section', 'g11_general_average', 'g11_adviser', 'g11_adviserContact',
            '',  # Blank column
            # Grade 12
            'g12_school', 'g12_schoolYear', 'g12_section', 'g12_general_average', 'g12_adviser', 'g12_adviserContact',
        ]

        for row_num, student in enumerate(students, start_row):
            for col_num, attribute in enumerate(excel_columns_order, start_column):
                col_letter = get_column_letter(col_num)

                if attribute == '':
                    # Skip blank columns
                    continue

                field_value = getattr(student, attribute, None)

                if attribute in ['health_bmi', 'general_average']:
                    if field_value is not None:
                        field_value = float(field_value)
                elif attribute == 'birthday':
                    # Export date in MM-DD-YYYY format
                    field_value = student.birthday.strftime('%m-%d-%Y') if student.birthday else None
                    sheet[f"{col_letter}{row_num}"] = field_value
                    sheet[f"{col_letter}{row_num}"].number_format = 'MM-DD-YYYY'
                else:
                    if field_value is not None:
                        field_value = str(field_value)
                    else:
                        field_value = ""

                sheet[f"{col_letter}{row_num}"] = field_value

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students_data_updated.xlsx'
        existing_wb.save(response)

        messages.success(request, "Data successfully exported to Excel!")

        return response

    except Exception as e:
        print(f"Error exporting data to Excel: {str(e)}")
        messages.error(request, "An error occurred while exporting data to Excel. Please try again.")

    return render(request, 'view_students.html')

@allowed_users(allowed_roles=['TEACHER'])
def export_classroom_students_to_excel(request):
    try:
        user = request.user
        teacher = get_object_or_404(Teacher, user=user)
        classroom = get_object_or_404(Classroom, teacher=teacher)

        students = Student.objects.filter(classroom=classroom)

        existing_wb = load_workbook('data/static/media/VRCNHS_STUDENT_TEMPLATE.xlsx')
        sheet = existing_wb.active

        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

        start_row = 2
        start_column = 2

        # Define a list representing the order of columns in the Excel file
        excel_columns_order = [
            'LRN', 'last_name', 'first_name', 'middle_name', 'suffix_name', 'status', 'birthday',
            'religion', 'other_religion', 'strand', 'age', 'sem', 'classroom', 'gradelevel', 'sex',
            'birth_place', 'mother_tongue', 'address', 'father_name', 'father_contact', 'mother_name',
            'mother_contact', 'guardian_name', 'guardian_contact', 'transfer_status', 'household_income',
            'is_returnee', 'is_a_dropout', 'is_a_working_student', 'health_bmi', 'general_average',
            'is_a_four_ps_scholar', 'notes',
            '',  # Blank column
            '',
            # Grade 7
            'g7_school', 'g7_schoolYear', 'g7_section', 'g7_general_average', 'g7_adviser', 'g7_adviserContact',
            '',  # Blank column
            # Grade 8
            'g8_school', 'g8_schoolYear', 'g8_section', 'g8_general_average', 'g8_adviser', 'g8_adviserContact',
            '',  # Blank column
            # Grade 9
            'g9_school', 'g9_schoolYear', 'g9_section', 'g9_general_average', 'g9_adviser', 'g9_adviserContact',
            '',  # Blank column
            # Grade 10
            'g10_school', 'g10_schoolYear', 'g10_section', 'g10_general_average', 'g10_adviser', 'g10_adviserContact',
            '',  # Blank column
            # Grade 11
            'g11_school', 'g11_schoolYear', 'g11_section', 'g11_general_average', 'g11_adviser', 'g11_adviserContact',
            '',  # Blank column
            # Grade 12
            'g12_school', 'g12_schoolYear', 'g12_section', 'g12_general_average', 'g12_adviser', 'g12_adviserContact',
        ]

        for row_num, student in enumerate(students, start_row):
            for col_num, attribute in enumerate(excel_columns_order, start_column):
                col_letter = get_column_letter(col_num)

                if attribute == '':
                    # Skip blank columns
                    continue

                if attribute == 'classroom':
                    # Extract the relevant information from the Classroom object
                    field_value = getattr(student.classroom, 'classroom', '')
                elif attribute == 'gradelevel':
                    # Extract the relevant information from the Gradelevel object
                    field_value = getattr(student.gradelevel, 'grade', '')
                else:
                    field_value = getattr(student, attribute, '')

                # Your existing logic for handling different field types

                sheet[f"{col_letter}{row_num}"] = field_value

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=classroom_students_data.xlsx'
        existing_wb.save(response)

        messages.success(request, "Classroom data successfully exported to Excel!")

        return response

    except Exception as e:
        print(f"Error exporting classroom data to Excel: {str(e)}")
        messages.error(request, "An error occurred while exporting classroom data to Excel. Please try again.")

    return render(request, 'user_page.html')

#EXPORT classroom
@allowed_users(allowed_roles=['ADMIN'])
def export_classrooms_to_excel(request):
    classrooms = Classroom.objects.all().values()
    df = pd.DataFrame(list(classrooms))

    # Convert DataFrame to Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="classrooms_data.xlsx"'

    with pd.ExcelWriter(response) as writer:
        df.to_excel(writer, index=False)

    return response

#IMPORT
#This view function handles student data import from an Excel file.
from datetime import datetime
def import_students_from_excel(request):
    # Processes the student import request.

    if request.method == 'POST':
        student_resource = StudentResource()
        dataset = Dataset()
        new_student = request.FILES['myfile']

        if not new_student.name.endswith('xlsx'):
            messages.error(request, 'Please upload an Excel file only (.xlsx)')
            # Check user group and redirect accordingly
            user = request.user
            if Group.objects.get(name='TEACHER') in user.groups.all():
                # Redirect to teacher's user_page.html
                return redirect('user_page')  # Replace 'teacher_user_page' with your actual URL or view name
            else:
                # Redirect to view_students page after successful import
                return redirect('students')

        try:
            imported_data = dataset.load(new_student.read(), format='xlsx')
            successfully_imported = 0

            for i, data in enumerate(imported_data, start=1):
                LRN_value = str(data[1]).strip()  # Convert to string and then strip

                # Check if LRN is empty or "None"
                if LRN_value is None or not LRN_value.strip():
                    print(f"Row {i}: LRN is empty or None. Stopping further processing.")
                    break  # Stop processing further students

                classroom_identifier = data[13]#H
                gradelevel_identifier = data[14]#J

                # Check if classroom_identifier is None
                if classroom_identifier is None:
                    print(f"Row {i}: Classroom Identifier is None. Stopping further processing.")
                    break  # Stop processing further students

                if gradelevel_identifier is None:
                    print(f"Row {i}: Gradelevel Identifier is None. Stopping further processing.")
                    break  # Stop processing further students

                try:
                    classroom_instance = Classroom.objects.get(classroom=classroom_identifier)
                    gradelevel_instance = Gradelevel.objects.get(grade=gradelevel_identifier)

                    # Check if the birthday is already a datetime object or a string
                    if isinstance(data[7], datetime):
                        converted_bday = data[7]
                    elif data[7] and isinstance(data[7], str):
                        try:
                            converted_bday = date_parser.parse(data[7])
                        except ValueError:
                            # If parsing fails, handle it as needed
                            converted_bday = None
                    else:
                        converted_bday = None


                    print(f"Row {i}: Classroom Instance:", classroom_instance)
                    print(f"Row {i}: Gradelevel Instance:", gradelevel_instance)

                    value = Student(
                        LRN=data[1],#B
                        last_name=data[2],#C
                        first_name=data[3], #D
                        middle_name=data[4],#E
                        suffix_name=data[5],#F
                        status=data[6],#G
                        birthday=converted_bday,  # Use the provided date without conversion
                        religion=data[8], #I
                        other_religion=data[9], #J
                        strand=data[10], #K
                        age=data[11],#L
                        sem=data[12], #M
                        classroom=classroom_instance, #N
                        gradelevel=gradelevel_instance, #0
                        sex=data[15], #P
                        birth_place=data[16], #Q
                        mother_tongue=data[17], #R
                        address=data[18], #S
                        father_name=data[19], #T
                        father_contact=data[20],#U
                        mother_name=data[21],#V
                        mother_contact=data[22],#W
                        guardian_name=data[23],#X
                        guardian_contact=data[24],#Y
                        transfer_status=data[25],
                        household_income=data[26], #Z
                        is_returnee=data[27],#AA
                        is_a_dropout=data[28], #AB
                        is_a_working_student=data[29],#AC
                        health_bmi=data[30],#AD
                        general_average=data[31], #AE
                        is_a_four_ps_scholar=data[32], #AF
                        notes=data[33], #AG

                        #Grade 7
                        g7_school = data[36], #AJ
                        g7_schoolYear = data[37],
                        g7_section = data[38],
                        g7_general_average = data[39],
                        g7_adviser = data[40],
                        g7_adviserContact = data[41],

                        #Grade 8
                        g8_school = data[43], #AJ
                        g8_schoolYear = data[44],
                        g8_section = data[45],
                        g8_general_average = data[46],
                        g8_adviser = data[47],
                        g8_adviserContact = data[48],

                        #Grade 9
                        g9_school = data[50], #AJ
                        g9_schoolYear = data[51],
                        g9_section = data[52],
                        g9_general_average = data[53],
                        g9_adviser = data[54],
                        g9_adviserContact = data[55],


                        #Grade 10
                        g10_school = data[57], 
                        g10_schoolYear = data[58],
                        g10_section = data[59],
                        g10_general_average = data[60],
                        g10_adviser = data[61],
                        g10_adviserContact = data[62],


                        #Grade 11
                        g11_school = data[64], #AJ
                        g11_schoolYear = data[65],
                        g11_section = data[66],
                        g11_general_average = data[67],
                        g11_adviser = data[68],
                        g11_adviserContact = data[69],


                        #Grade 12
                        g12_school = data[71], #AJ
                        g12_schoolYear = data[72],
                        g12_section = data[73],
                        g12_general_average = data[74],
                        g12_adviser = data[75],
                        g12_adviserContact = data[76],

                    )
                    value.save()
                    successfully_imported += 1
                    print(f"Row {i}: Successfully imported student:", value)

                except Exception as e:
                    error_message = f"Row {i}: Error saving student data: {str(e)}"
                    print(error_message)
                    messages.error(request, error_message)
                    break

            if successfully_imported > 0:
                messages.success(request, f"Successfully imported {successfully_imported} student(s) into the database.")

        except Exception as e:
            print(f"Error loading student/s from the file: {str(e)}")

        # Check user group and redirect accordingly
        user = request.user
        if Group.objects.get(name='TEACHER') in user.groups.all():
            # Redirect to teacher's user_page.html
            return redirect('user_page')  # Replace 'teacher_user_page' with your actual URL or view name
        else:
            # Redirect to view_students page after successful import
            return redirect('students')

    return render(request, 'view_students.html')


def import_classrooms_from_excel(request):
    # Processes the classroom import request.

    if request.method == 'POST':  # Check if a POST request (file submission) was made
        classroom_resource = ClassroomResource()  # Instantiate a ClassroomResource object (likely for data validation)
        dataset = Dataset()  # Instantiate a Dataset object for loading the Excel data
        new_classroom = request.FILES['myfile']  # Retrieve the uploaded file from the request

        if not new_classroom.name.endswith('xlsx'):  # Validate file extension
            messages.error(request, 'Please upload an Excel file only (.xlsx)')  # Display an error message
            return render(request, 'grade_sections.html')  # Re-render the view_classrooms.html page

        try:
            imported_data = dataset.load(new_classroom.read(), format='xlsx')  # Load Excel data into a dataset
            successfully_imported = 0

            for data in imported_data:  # Iterate through each classroom record in the dataset
                grade_level_identifier = data[0]  # Extract grade level identifier from the 1st column (index 0)
                grade_level_instance = Gradelevel.objects.get(grade=grade_level_identifier)  # Retrieve grade level object

                try:
                    # Create a new Classroom object with data from the Excel row
                    classroom = Classroom(
                        gradelevel=grade_level_instance,  # Assign retrieved grade level object
                        classroom=data[1],
                        teacher=None  # Replace None with actual teacher data if available
                    )
                    classroom.save()  # Save the classroom object to the database
                    successfully_imported += 1
                except Exception as e:
                    messages.error(request, f"Error saving classroom data: {str(e)}")

            if successfully_imported > 0:
                messages.success(request, f"Successfully imported {successfully_imported} classroom(s) into the database.")

        except Exception as e:
            messages.info(request, f"Error loading classroom/s from the file: {str(e)}")

    return render(request, 'grade_sections.html')  # Re-render the view_classrooms.html page after processing


#BULKPROMOTE

@login_required(login_url='login')
def students_for_promotion(request):
    # Filter students who are either 'For Promotion' or 'For Retention' and are in the "SECTIONING" classroom
    query_conditions = Q(
        (Q(status='For Promotion') | Q(status='For Retention')) &
        Q(classroom__classroom='SECTIONING')
    )

    # Apply these conditions to each grade level query
    students_grade_8 = Student.objects.filter(
        query_conditions,
        gradelevel__grade='Grade 8',
    )

    students_grade_9 = Student.objects.filter(
        query_conditions,
        gradelevel__grade='Grade 9',
    )

    students_grade_10 = Student.objects.filter(
        query_conditions,
        gradelevel__grade='Grade 10',
    )

    students_grade_11 = Student.objects.filter(
        query_conditions,
        gradelevel__grade='Grade 11',
    )

    students_grade_12 = Student.objects.filter(
        query_conditions,
        gradelevel__grade='Grade 12',
    )

    # Filter students for departure based on status
    for_departure = Student.objects.filter(
        Q(status='For Graduation') | Q(status='For Dropout/Transfer')
    )

    classrooms = Classroom.objects.all()

    # Assuming you have a predefined list of grade levels
    grade_levels = Gradelevel.objects.all()
    grade_level_classrooms = {grade.id: Classroom.objects.filter(gradelevel=grade) for grade in grade_levels}

    for grade, students in [('Grade 8', students_grade_8), ('Grade 9', students_grade_9), ('Grade 10', students_grade_10), ('Grade 11', students_grade_11), ('Grade 12', students_grade_12)]:
        for student in students:
            student.classroom_options = grade_level_classrooms.get(student.gradelevel_id, [])
            student.LRN_str = str(student.LRN)
            print(student.LRN, student.classroom_options)  # Debugging line to check values

    # Prepare context data with the updated queries
    context = {
        'students_grade_8': students_grade_8,
        'students_grade_9': students_grade_9,
        'students_grade_10': students_grade_10,
        'students_grade_11': students_grade_11,
        'students_grade_12': students_grade_12,
        'classrooms': classrooms,
        'for_departure': for_departure,
    }

    return render(request, 'students_for_promotion.html', context)

@login_required(login_url='login')
def assign_classroom_bulk(request, grade):
    if request.method == 'POST':
        # List to store student LRNs for updating status
        student_lrns = []

        for key, value in request.POST.items():
            if key.startswith('classroom_'):
                student_lrn_str = key.split('_')[1] 
                classroom_id = value

                try:
                    student = Student.objects.get(LRN=str(student_lrn_str))  # Convert LRN to string for comparison
                    classroom = Classroom.objects.get(id=classroom_id)
                    student.classroom = classroom
                    student.status = 'Currently Enrolled'  # Update status to "Processing"
                    student.general_average = None
                    student.save()
                    student_lrns.append(student_lrn_str)  # Use LRN_str here

                except Student.DoesNotExist:
                    messages.warning(request, f"Student with LRN {student_lrn_str} does not exist. Skipping.")
                    continue

                except Classroom.DoesNotExist:
                    messages.warning(request, f"Classroom with ID {classroom_id} does not exist. Skipping.")
                    continue

        # Update status for all selected students
        Student.objects.filter(LRN__in=student_lrns).update(status='Currently Enrolled')

        messages.success(request, 'Students assigned to classrooms successfully.')
        return redirect('students_for_promotion')  # Redirect back to the promotion page
    
@login_required(login_url='login')
def bulk_promote_students(request):
    response_data = {'success': False, 'message': 'Failed to promote students'}

    # Check if there are students with the status "Currently Enrolled"
    if Student.objects.filter(classroom__teacher__user=request.user, status='Currently Enrolled').exists():
        messages.error(request, 'No student should have a status of "Currently Enrolled" anymore if you want to Promote in Bulk')
        return redirect('user_page')

    if request.method == 'POST':
        try:    
            try:
                # Assuming there is a direct reference from Classroom to Teacher
                user_classroom = Classroom.objects.get(teacher__user=request.user)
            except Classroom.DoesNotExist:
                # Handle the case where there is no associated Classroom for the teacher
                messages.error(request, 'User is not associated with a classroom.')
                return redirect('students')  # Replace 'students' with the actual URL name of your students page

            # Determine the next grade level based on the teacher's current grade level
            current_grade = user_classroom.gradelevel.grade
            next_grade = get_next_grade(current_grade)

            if next_grade is None:
                messages.error(request, 'Unable to determine the next grade level.')
                return redirect('students')  # Replace 'students' with the actual URL name of your students page

            next_grade_instance = Gradelevel.objects.get(grade=next_grade)

            students_to_promote = Student.objects.filter(classroom=user_classroom)

            for student in students_to_promote:
                # Store current classroom name as the previous section before changing it
                student.previous_section = str(student.classroom)
                
                if student.status == 'For Promotion':
                    # Specific cases based on the current grade
                    if current_grade == 'Grade 7':
                        student.g7_section = user_classroom.classroom
                        student.g7_general_average = student.general_average
                        student.g7_adviser = f"{request.user.first_name} {request.user.last_name}"
                    elif current_grade == 'Grade 8':
                        student.g8_section = user_classroom.classroom
                        student.g8_general_average = student.general_average
                        student.g8_adviser = f"{request.user.first_name} {request.user.last_name}"
                    elif current_grade == 'Grade 9':
                        student.g9_section = user_classroom.classroom
                        student.g9_general_average = student.general_average
                        student.g9_adviser = f"{request.user.first_name} {request.user.last_name}"
                    elif current_grade == 'Grade 10':
                        student.g10_section = user_classroom.classroom
                        student.g10_general_average = student.general_average
                        student.g10_adviser = f"{request.user.first_name} {request.user.last_name}"
                    elif current_grade == 'Grade 11':
                        student.g11_section = user_classroom.classroom
                        student.g11_general_average = student.general_average
                        student.g11_adviser = f"{request.user.first_name} {request.user.last_name}"
                    elif current_grade == 'Grade 12':
                        student.g12_section = user_classroom.classroom
                        student.g12_general_average = student.general_average
                        student.g12_adviser = f"{request.user.first_name} {request.user.last_name}"
                    student.gradelevel = next_grade_instance
                    student.classroom = Classroom.objects.get(classroom='SECTIONING')

                elif student.status == 'For Retention':
                    # Specific cases based on the current grade
                    if current_grade == 'Grade 7':
                        student.g7_section = user_classroom.classroom
                        student.g7_general_average = student.general_average
                        student.g7_adviser = f"{request.user.first_name} {request.user.last_name}"
                    elif current_grade == 'Grade 8':
                        student.g8_section = user_classroom.classroom
                        student.g8_general_average = student.general_average
                        student.g8_adviser = f"{request.user.first_name} {request.user.last_name}"
                    elif current_grade == 'Grade 9':
                        student.g9_section = user_classroom.classroom
                        student.g9_general_average = student.general_average
                        student.g9_adviser = f"{request.user.first_name} {request.user.last_name}"
                    elif current_grade == 'Grade 10':
                        student.g10_section = user_classroom.classroom
                        student.g10_general_average = student.general_average
                        student.g10_adviser = f"{request.user.first_name} {request.user.last_name}"
                    elif current_grade == 'Grade 11':
                        student.g11_section = user_classroom.classroom
                        student.g11_general_average = student.general_average
                        student.g11_adviser = f"{request.user.first_name} {request.user.last_name}"
                    elif current_grade == 'Grade 12':
                        student.g12_section = user_classroom.classroom
                        student.g12_general_average = student.general_average
                        student.g12_adviser = f"{request.user.first_name} {request.user.last_name}"
                    student.classroom = Classroom.objects.get(classroom='SECTIONING')
                elif student.status in ['For Graduation', 'For Dropout/Transfer']:
                    student.classroom = Classroom.objects.get(classroom='FOR DEPARTURE')
                student.save()
                

            response_data['success'] = True
            response_data['message'] = f'Bulk promotion to {next_grade} successful!'
            messages.success(request, response_data['message'])

            return JsonResponse(response_data)
        except Exception as e:
            print(e)
            response_data['message'] = 'Failed to promote students'
            messages.error(request, response_data['message'])
            return JsonResponse(response_data)

    
    messages.success(request, 'Students Promoted! Please Refresh')
    return redirect('user_page')


def get_next_grade(current_grade):
    # Define the grade progression logic here
    grade_sequence = ['Grade 7', 'Grade 8', 'Grade 9', 'Grade 10', 'Grade 11', 'Grade 12']

    try:
        current_index = grade_sequence.index(current_grade)
        next_index = current_index + 1

        if next_index < len(grade_sequence):
            return grade_sequence[next_index]
        else:
            return None  # No next grade (end of sequence)
    except ValueError:
        return None  # Current grade not found in the sequence
    
@allowed_users(allowed_roles=['ADMIN', 'TEACHER'])
def export_and_delete_students_for_departure(request):
    try:
        # Retrieve students for departure
        students_for_departure = Student.objects.filter(
            Q(status='For Graduation') | Q(status='For Dropout/Transfer'),
        )

        # Load the template workbook
        template_path = 'data/static/media/VRCNHS_STUDENT_TEMPLATE.xlsx'
        existing_wb = load_workbook(template_path)
        sheet = existing_wb.active

        # Clear existing data
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

        start_row = 2
        start_column = 2

        date_style = NamedStyle(name='date_style', number_format='MM-DD-YYYY')

        # Define a list representing the order of columns in the Excel file
        excel_columns_order = [
            'LRN', 'last_name', 'first_name', 'middle_name', 'suffix_name', 'status', 'birthday',
            'religion', 'other_religion', 'strand', 'age', 'sem', 'classroom', 'gradelevel', 'sex',
            'birth_place', 'mother_tongue', 'address', 'father_name', 'father_contact', 'mother_name',
            'mother_contact', 'guardian_name', 'guardian_contact', 'transfer_status', 'household_income',
            'is_returnee', 'is_a_dropout', 'is_a_working_student', 'health_bmi', 'general_average',
            'is_a_four_ps_scholar', 'notes',
            '',  # Blank column
            '',
            # Grade 7
            'g7_school', 'g7_schoolYear', 'g7_section', 'g7_general_average', 'g7_adviser', 'g7_adviserContact',
            '',  # Blank column
            # Grade 8
            'g8_school', 'g8_schoolYear', 'g8_section', 'g8_general_average', 'g8_adviser', 'g8_adviserContact',
            '',  # Blank column
            # Grade 9
            'g9_school', 'g9_schoolYear', 'g9_section', 'g9_general_average', 'g9_adviser', 'g9_adviserContact',
            '',  # Blank column
            # Grade 10
            'g10_school', 'g10_schoolYear', 'g10_section', 'g10_general_average', 'g10_adviser', 'g10_adviserContact',
            '',  # Blank column
            # Grade 11
            'g11_school', 'g11_schoolYear', 'g11_section', 'g11_general_average', 'g11_adviser', 'g11_adviserContact',
            '',  # Blank column
            # Grade 12
            'g12_school', 'g12_schoolYear', 'g12_section', 'g12_general_average', 'g12_adviser', 'g12_adviserContact',
        ]

        for row_num, student in enumerate(students_for_departure, start_row):
            for col_num, attribute in enumerate(excel_columns_order, start_column):
                col_letter = get_column_letter(col_num)

                if attribute == '':
                    # Skip blank columns
                    continue

                field_value = getattr(student, attribute, None)

                if attribute in ['LRN', 'age', 'father_contact', 'mother_contact', 'guardian_contact', 'adviser_contact']:
                    if field_value is not None:
                        field_value = int(field_value)
                elif attribute in ['health_bmi', 'general_average']:
                    if field_value is not None:
                        field_value = float(field_value)
                elif attribute == 'birthday':
                    # Export date in MM-DD-YYYY format
                    field_value = student.birthday.strftime('%m-%d-%Y') if student.birthday else None
                    sheet[f"{col_letter}{row_num}"] = field_value
                    sheet[f"{col_letter}{row_num}"].number_format = 'MM-DD-YYYY'
                elif attribute == 'last_grade_level':
                    # Check if last_grade_level is a string with non-numeric characters
                    if field_value is not None and any(c.isalpha() for c in field_value):
                        # If it contains non-numeric characters, export as a string
                        field_value = str(field_value)
                    else:
                        # Otherwise, try converting to integer
                        try:
                            field_value = int(field_value)
                        except (ValueError, TypeError):
                            print(f"Error converting last_grade_level to int: {field_value}")
                            field_value = None
                else:
                    if field_value is not None:
                        field_value = str(field_value)
                    else:
                        field_value = ""

                sheet[f"{col_letter}{row_num}"] = field_value

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students_for_departure.xlsx'
        existing_wb.save(response)

        # Delete students after successful export
        students_for_departure.delete()

        messages.success(request, "Data successfully exported to Excel and students deleted!")

        return response

    except Exception as e:
        print(f"Error exporting data to Excel: {str(e)}")
        messages.error(request, "An error occurred while exporting data to Excel. Please try again.")

    return render(request, 'view_students.html')

